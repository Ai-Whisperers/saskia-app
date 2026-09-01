"""Tests/conftest.py — shared pytest fixtures for the Saskia RMS test suite.

Per docs/operations/2026-09-fase-1-specs.md §9 (test-suite minimum).

Key principle: tests must NEVER write to the production DB path
(`~/.local/share/AIW-Saskia/rms.sqlite`). The `tmp_db_path` fixture
forces every test to use a temp directory.
"""

from __future__ import annotations

import pytest


@pytest.fixture(autouse=True)
def tmp_db_path(tmp_path, monkeypatch):
    """Force every test to use a fresh temp DB.

    Sets AIW_SASKIA_DB_PATH and AIW_SASKIA_DATA_DIR/BACKUP_DIR/LOG_DIR
    to tmp_path. autouse=True means every test gets this isolation.
    """
    monkeypatch.setenv("AIW_SASKIA_DB_PATH", str(tmp_path / "test.sqlite"))
    monkeypatch.setenv("AIW_SASKIA_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setenv("AIW_SASKIA_BACKUP_DIR", str(tmp_path / "backups"))
    monkeypatch.setenv("AIW_SASKIA_LOG_DIR", str(tmp_path / "logs"))
    monkeypatch.setenv("BIND_HOST", "127.0.0.1")
    monkeypatch.setenv("PORT", "8765")
    return tmp_path


@pytest.fixture
def temp_dir(tmp_path):
    """A fresh temp directory for file-based tests."""
    return tmp_path


@pytest.fixture
def make_decimal():
    """Factory for Decimal values; convenient for parametrize."""
    from decimal import Decimal

    def _make(value):
        return Decimal(str(value))

    return _make


@pytest.fixture
def app_engine(tmp_db_path):
    """Create a file-based SQLite engine in tmp_path with WAL pragmas.

    Use this fixture when you need a DB but not a full FastAPI app.

    File-based (not :memory:) so all connections share the same DB.
    SQLite's :memory: creates a separate DB per connection unless you
    pin a StaticPool — but our WAL pragma listener then can't safely
    apply journal_mode to per-connection in-memory DBs. File-based
    solves both problems.
    """
    from app.rms.db import init_db, make_engine

    engine = make_engine(f"sqlite:///{tmp_db_path}/test.sqlite")
    init_db(engine)
    return engine


@pytest.fixture
def session_factory(app_engine):
    """sessionmaker bound to the app_engine fixture."""
    from app.rms.db import make_session_factory

    return make_session_factory(app_engine)


@pytest.fixture(autouse=True)
def reset_app_state():
    """Reset FastAPI app.state between tests so the lifespan override is clean.

    Without this, app.state.engine / app.state.session_factory from one test
    leak into the next, causing tests that should be isolated to share state.
    """
    from app.rms.main import app

    # Clear any state set by previous tests; the `client` fixture repopulates.
    if hasattr(app.state, "engine"):
        del app.state.engine
    if hasattr(app.state, "session_factory"):
        del app.state.session_factory
    yield


@pytest.fixture
def client(session_factory, monkeypatch):
    """TestClient wired with a session_factory that uses app_engine.

    Replaces make_engine_dialect in app.rms.main with a deterministic version
    that returns the test engine. This prevents the lifespan from creating
    a separate engine (and a separate DB) at startup.
    """
    from app.rms import main as main_module

    test_engine = session_factory.kw["bind"]

    def _make_engine_for_test(url=None, *, for_tests=False):
        # Always return our test engine, ignore URL or args.
        return test_engine

    monkeypatch.setattr(main_module, "make_engine_dialect", _make_engine_for_test)

    from fastapi.testclient import TestClient

    with TestClient(main_module.app) as c:
        yield c


# --- xlsx fixture: synthetic HEREBUS workbook for import/export tests ---


@pytest.fixture
def mini_xlsx_path(tmp_path):
    """Return a Path to a synthetic HEREBUS-format .xlsx fixture.

    3 ingredients, 1 recipe (Muffin), 3 ingredient lines, 2 products.
    Sales + StockMoves sheets are present but empty (they're runtime data).

    Generated on demand; never checked into git.
    """
    from pathlib import Path

    from openpyxl import Workbook

    path: Path = tmp_path / "mini.xlsx"
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    ws = wb.create_sheet("Ingredientes")
    ws.append(["id", "name", "unit", "stock_qty", "purchase_price_gs", "min_stock_qty", "notes"])
    ws.append([1, "Harina", "kg", 2.0, "Gs. 5.000", 0.5, None])
    ws.append([2, "Azúcar", "kg", 1.5, "Gs. 4.000", 0.3, None])
    ws.append([3, "Huevo", "und", 20.0, "Gs. 1.500", 6.0, None])

    ws = wb.create_sheet("Recetas")
    ws.append(["id", "name", "yield_qty", "yield_unit", "notes"])
    ws.append([1, "Muffin", 12.0, "und", None])

    ws = wb.create_sheet("Lineas")
    ws.append(
        [
            "id",
            "recipe_id",
            "recipe_name",
            "line_kind",
            "line_ref_id",
            "target_name",
            "qty",
            "notes",
        ]
    )
    ws.append([1, 1, "Muffin", "ingredient", 1, "Harina", 0.3, None])
    ws.append([2, 1, "Muffin", "ingredient", 2, "Azúcar", 0.2, None])
    ws.append([3, 1, "Muffin", "ingredient", 3, "Huevo", 2.0, None])

    ws = wb.create_sheet("Productos")
    ws.append(
        [
            "id",
            "name",
            "portion_label",
            "sale_price_gs",
            "recipe_id",
            "recipe_name",
            "notes",
        ]
    )
    ws.append([1, "Muffin", "1 muffin", "Gs. 8.000", 1, "Muffin", None])
    ws.append([2, "Mystery", "1 unidad", "Gs. 5.000", None, None, "no recipe"])

    ws = wb.create_sheet("Ventas")
    ws.append(
        [
            "id",
            "sold_at",
            "product_id",
            "product_name",
            "qty",
            "unit_price_gs",
            "notes",
            "voided_at",
        ]
    )

    ws = wb.create_sheet("StockMoves")
    ws.append(["id", "sale_id", "affected_recipe_id", "ingredient_id", "qty_delta"])

    wb.save(str(path))
    return path

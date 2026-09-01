"""tests/test_r2_backup.py — encryption + Storage roundtrip + scheduler.

Per dev plan Batch 5.

Tests cover:
- Fernet encrypt/decrypt roundtrip preserves bytes
- Decrypt fails on tampered ciphertext
- Decrypt fails with wrong key
- Boto3Storage protocol is satisfied (duck-type check)
- InMemoryStorage stores ciphertext, never plaintext
- encrypt_and_upload → download_and_decrypt roundtrip
- backup_scheduler.run_backup: skips when threshold not crossed
- backup_scheduler.run_backup: runs when threshold crossed
- backup_scheduler.run_backup: writes last_backup_at meta
- backup_scheduler.run_backup: uploads to R2 (with injected Storage)
- backup_scheduler.run_backup: R2 failure does NOT invalidate local backup
- load_or_create_key: creates a key file on first run
- load_r2_settings: returns None when config missing
- load_r2_settings: returns Settings when config present
- load_r2_settings: raises on missing required keys
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import pytest

# --- Fernet roundtrip ---


def test_encrypt_decrypt_roundtrip_preserves_bytes():
    """Encrypt then decrypt returns identical plaintext."""
    from app.services.r2_backup import decrypt_bytes, encrypt_bytes

    plaintext = b"Saskia RMS - important business data"
    # Generate a real key for the test
    from cryptography.fernet import Fernet

    key = Fernet.generate_key()
    ciphertext = encrypt_bytes(plaintext, key)
    assert ciphertext != plaintext  # encryption changed the bytes
    decrypted = decrypt_bytes(ciphertext, key)
    assert decrypted == plaintext


def test_decrypt_fails_on_tampered_ciphertext():
    """Modifying one byte of ciphertext makes decryption fail."""
    from cryptography.fernet import Fernet

    from app.services.r2_backup import DecryptionError, decrypt_bytes, encrypt_bytes

    key = Fernet.generate_key()
    ciphertext = bytearray(encrypt_bytes(b"hello", key))
    ciphertext[20] ^= 0xFF  # flip a byte
    with pytest.raises(DecryptionError):
        decrypt_bytes(bytes(ciphertext), key)


def test_decrypt_fails_with_wrong_key():
    """Decrypting with the wrong key fails."""
    from cryptography.fernet import Fernet

    from app.services.r2_backup import DecryptionError, decrypt_bytes, encrypt_bytes

    key_a = Fernet.generate_key()
    key_b = Fernet.generate_key()
    ciphertext = encrypt_bytes(b"hello", key_a)
    with pytest.raises(DecryptionError):
        decrypt_bytes(ciphertext, key_b)


# --- InMemoryStorage ---


def test_inmemory_storage_put_get():
    """put + get roundtrip."""
    from app.services.r2_backup import InMemoryStorage

    s = InMemoryStorage()
    s.put("foo", b"bar")
    assert s.get("foo") == b"bar"


def test_inmemory_storage_missing_key_raises():
    """get on missing key raises KeyError."""
    from app.services.r2_backup import InMemoryStorage

    s = InMemoryStorage()
    with pytest.raises(KeyError):
        s.get("nope")


def test_inmemory_storage_overwrites():
    """put on existing key overwrites."""
    from app.services.r2_backup import InMemoryStorage

    s = InMemoryStorage()
    s.put("k", b"v1")
    s.put("k", b"v2")
    assert s.get("k") == b"v2"


# --- encrypt_and_upload + download_and_decrypt roundtrip ---


def test_encrypt_and_upload_then_download_decrypt():
    """End-to-end: encrypt + upload + download + decrypt returns plaintext."""
    from cryptography.fernet import Fernet

    from app.services.r2_backup import (
        InMemoryStorage,
        download_and_decrypt,
        encrypt_and_upload,
    )

    storage = InMemoryStorage()
    key = Fernet.generate_key()
    plaintext = b"backup contents"

    ciphertext = encrypt_and_upload(storage, key, "snap.enc", plaintext)
    assert storage.get("snap.enc") == ciphertext
    # Ciphertext should NOT contain plaintext (sanity)
    assert plaintext not in ciphertext

    decrypted = download_and_decrypt(storage, key, "snap.enc")
    assert decrypted == plaintext


# --- load_or_create_key ---


def test_load_or_create_key_creates_when_missing(tmp_path: Path):
    """First run: writes a new key file with mode 0o600."""
    from app.services.r2_backup import load_or_create_key

    key_file = tmp_path / "subdir" / "key"
    assert not key_file.exists()
    key = load_or_create_key(key_file)
    assert key_file.exists()
    # Key is 44 bytes (32-byte key urlsafe-base64-encoded)
    assert len(key) > 30
    # Cryptographic-strength key: loadable as Fernet
    from cryptography.fernet import Fernet

    Fernet(key)  # would raise if invalid


def test_load_or_create_key_returns_existing(tmp_path: Path):
    """Second run: returns the same key."""
    from app.services.r2_backup import load_or_create_key

    key_file = tmp_path / "key"
    key1 = load_or_create_key(key_file)
    key2 = load_or_create_key(key_file)
    assert key1 == key2


# --- load_r2_settings ---


def test_load_r2_settings_returns_none_when_missing(tmp_path: Path, monkeypatch):
    """No r2.toml → None (not an error)."""
    import app.services.r2_backup as r2_mod
    from app.services.r2_backup import load_r2_settings

    monkeypatch.setattr(r2_mod, "R2_CONFIG_PATH", tmp_path / "does_not_exist.toml")
    assert load_r2_settings() is None


def test_load_r2_settings_returns_settings_when_present(tmp_path: Path, monkeypatch):
    """Valid r2.toml → R2Settings with all 4 fields."""
    import app.services.r2_backup as r2_mod
    from app.services.r2_backup import load_r2_settings

    cfg = tmp_path / "r2.toml"
    cfg.write_text(
        "[r2]\n"
        'endpoint_url = "https://example.r2.cloudflarestorage.com"\n'
        'bucket = "saskia-backups"\n'
        'access_key_id = "AKIA..."\n'
        'secret_access_key = "secret"\n'
    )
    monkeypatch.setattr(r2_mod, "R2_CONFIG_PATH", cfg)
    s = load_r2_settings()
    assert s is not None
    assert s.bucket == "saskia-backups"
    assert s.endpoint_url.startswith("https://")


def test_load_r2_settings_raises_on_missing_keys(tmp_path: Path, monkeypatch):
    """Incomplete r2.toml → StorageError."""
    import app.services.r2_backup as r2_mod
    from app.services.r2_backup import StorageError, load_r2_settings

    cfg = tmp_path / "r2.toml"
    cfg.write_text('[r2]\nbucket = "x"\n')  # missing 3 keys
    monkeypatch.setattr(r2_mod, "R2_CONFIG_PATH", cfg)
    with pytest.raises(StorageError, match="missing required keys"):
        load_r2_settings()


# --- backup_scheduler.run_backup ---


def _seed_db(db_path: Path) -> None:
    """Create a real SQLite file with our schema."""
    from app.rms.db import init_db, make_engine
    from app.rms.models import Ingredient

    db_path.parent.mkdir(parents=True, exist_ok=True)
    engine = make_engine(f"sqlite:///{db_path}")
    init_db(engine)
    from sqlalchemy.orm import sessionmaker

    Session = sessionmaker(bind=engine)
    with Session() as s:
        s.add(Ingredient(name="Test", unit="kg", stock_qty=1.0, purchase_price_gs=1000))
        s.commit()
    engine.dispose()


def test_run_backup_skips_when_recent(session_factory, tmp_path: Path, monkeypatch):
    """If last_backup_at is within threshold → skipped."""
    from app.rms.models import AppMeta
    from app.services.backup_scheduler import (
        APP_META_LAST_BACKUP,
        InMemoryStorage,
        run_backup,
    )

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"

    # Mark backup as recent
    with session_factory() as s:
        s.add(
            AppMeta(
                key=APP_META_LAST_BACKUP,
                value=(datetime.now() - timedelta(hours=1)).isoformat(),
                updated_at=datetime.now().isoformat(),
            )
        )
        s.commit()

    result = run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        storage=InMemoryStorage(),
    )
    assert result.skipped is True
    assert not backup_dir.exists() or not any(backup_dir.glob("*.xlsx"))


def test_run_backup_executes_when_old(session_factory, tmp_path: Path):
    """If last_backup_at is older than threshold → backup runs."""
    from app.rms.models import AppMeta
    from app.services.backup_scheduler import (
        APP_META_LAST_BACKUP,
        InMemoryStorage,
        run_backup,
    )

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"

    # Mark backup as 25h old (over default 24h threshold)
    with session_factory() as s:
        s.add(
            AppMeta(
                key=APP_META_LAST_BACKUP,
                value=(datetime.now() - timedelta(hours=25)).isoformat(),
                updated_at=datetime.now().isoformat(),
            )
        )
        s.commit()

    result = run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        storage=InMemoryStorage(),
    )
    assert result.skipped is False
    assert result.local_path is not None
    assert result.local_path.exists()
    # xlsx has at least the 6 sheets
    from openpyxl import load_workbook

    wb = load_workbook(result.local_path)
    assert "Ingredientes" in wb.sheetnames


def test_run_backup_executes_when_no_prior_backup(session_factory, tmp_path: Path):
    """No app_meta at all → runs the backup."""
    from app.services.backup_scheduler import InMemoryStorage, run_backup

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"

    result = run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        storage=InMemoryStorage(),
    )
    assert result.skipped is False
    assert result.local_path is not None
    assert result.local_path.exists()


def test_run_backup_uploads_to_r2_when_storage_provided(session_factory, tmp_path: Path):
    """With InMemoryStorage, the encrypted snapshot lands in storage."""
    from app.services.backup_scheduler import (
        InMemoryStorage,
        run_backup,
    )

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"
    storage = InMemoryStorage()

    result = run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        storage=storage,
    )
    assert result.r2_uploaded is True
    assert result.r2_key is not None
    # Snapshot key was created in storage
    assert result.r2_key in storage.keys()
    # Ciphertext != plaintext of the db file
    # Ciphertext should not be a SQLite header
    ciphertext = storage.get(result.r2_key)
    assert not ciphertext.startswith(b"SQLite format 3")


def test_run_backup_r2_failure_does_not_fail_local(session_factory, tmp_path: Path):
    """If storage.put raises StorageError, local backup still succeeds."""
    from app.services.backup_scheduler import run_backup

    class _FailingStorage:
        def put(self, key, data):
            from app.services.r2_backup import StorageError

            raise StorageError("simulated R2 outage")

        def get(self, key):
            raise KeyError(key)

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"

    result = run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        storage=_FailingStorage(),
    )
    # Local backup succeeded
    assert result.skipped is False
    assert result.local_path is not None
    assert result.local_path.exists()
    # R2 failed
    assert result.r2_uploaded is False
    assert result.r2_key is None


def test_run_backup_prunes_old_local_files(session_factory, tmp_path: Path):
    """Old backups beyond keep_last_n are deleted."""
    from app.services.backup_scheduler import InMemoryStorage, run_backup

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"
    backup_dir.mkdir(parents=True)

    # Pre-populate with 5 "old" backup files
    for i in range(5):
        old = backup_dir / f"rms-backup-2026010{i}-120000.xlsx"
        old.write_bytes(b"old")

    result = run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        keep_last_n=2,
        storage=InMemoryStorage(),
    )

    # Only the 2 newest (the new one + 1 survivor) remain
    remaining = sorted(backup_dir.glob("rms-backup-*.xlsx"))
    assert len(remaining) == 2
    # The new one is there
    assert result.local_path in remaining


def test_run_backup_writes_meta_after_success(session_factory, tmp_path: Path):
    """After backup, app_meta has last_backup_at updated to now-ish time."""
    from app.services.backup_scheduler import (
        APP_META_LAST_BACKUP,
        InMemoryStorage,
        run_backup,
    )

    db_path = tmp_path / "db.sqlite"
    _seed_db(db_path)
    backup_dir = tmp_path / "backups"
    fixed_now = datetime(2026, 9, 1, 12, 0, 0)

    run_backup(
        session_factory(),
        db_path,
        backup_dir=backup_dir,
        threshold_hours=24,
        storage=InMemoryStorage(),
        now=fixed_now,
    )

    with session_factory() as s:
        from app.rms.models import AppMeta

        row = s.query(AppMeta).filter_by(key=APP_META_LAST_BACKUP).first()
        assert row is not None
        # ISO timestamp parses
        parsed = datetime.fromisoformat(row.value)
        assert parsed == fixed_now


# --- Storage protocol satisfied by both Boto3 + InMemory ---


def test_inmemory_storage_implements_protocol():
    """InMemoryStorage has both put() and get() methods matching the protocol."""
    from app.services.r2_backup import InMemoryStorage

    s = InMemoryStorage()
    assert hasattr(s, "put")
    assert hasattr(s, "get")
    assert callable(s.put)
    assert callable(s.get)

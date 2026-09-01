"""app/services/export_xlsx.py — SQLite → .xlsx (HEREBUS format).

Per dev plan §9 Task 6 + v2 §6 (Excel export).

Writes a symmetric mirror of import_xlsx: same 6 sheets, same column order.
Roundtrip-safe: import(export(X)) ≡ X for non-derived state.

Money discipline: integer Gs. values written via the canonical format
(Gs. 1.234.567) so the resulting file is human-readable in Excel.

What we DON'T write here:
- StockMoves (derived; see import_xlsx note on why we ignore them on read)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.rms.models import Ingredient, Product, Recipe, Sale
from app.rms.money import format_gs

# Sheet column definitions — single source of truth for import + export.
INGREDIENTES_COLS = [
    "id",
    "name",
    "unit",
    "stock_qty",
    "purchase_price_gs",
    "min_stock_qty",
    "notes",
]
RECETAS_COLS = ["id", "name", "yield_qty", "yield_unit", "notes"]
LINEAS_COLS = [
    "id",
    "recipe_id",
    "recipe_name",
    "line_kind",
    "line_ref_id",
    "target_name",
    "qty",
    "notes",
]
PRODUCTOS_COLS = [
    "id",
    "name",
    "portion_label",
    "sale_price_gs",
    "recipe_id",
    "recipe_name",
    "notes",
]
VENTAS_COLS = [
    "id",
    "sold_at",
    "product_id",
    "product_name",
    "qty",
    "unit_price_gs",
    "notes",
    "voided_at",
]
STOCKMOVES_COLS = [
    "id",
    "sale_id",
    "affected_recipe_id",
    "ingredient_id",
    "qty_delta",
]


def _write_header(ws, cols: list[str]) -> None:
    for i, col in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=col)


def _money_cell(value: int | None) -> str | None:
    """Format integer Gs. for export. None stays None (Excel sees empty)."""
    if value is None:
        return None
    return format_gs(value)


def _autosize(ws, max_width: int = 40) -> None:
    """Set column widths from content. Capped at max_width."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def to_file(session: Session, path: str | Path) -> Path:
    """Export current DB state to a HEREBUS-format .xlsx.

    Returns the absolute Path of the written file.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default sheet; we'll add named ones.
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    # --- Ingredientes ---
    ws = wb.create_sheet("Ingredientes")
    _write_header(ws, INGREDIENTES_COLS)
    for ing in session.scalars(select(Ingredient).order_by(Ingredient.id)).all():
        ws.append(
            [
                ing.id,
                ing.name,
                ing.unit,
                ing.stock_qty,
                _money_cell(ing.purchase_price_gs),
                ing.min_stock_qty,
                ing.notes,
            ]
        )
    _autosize(ws)

    # --- Recetas ---
    ws = wb.create_sheet("Recetas")
    _write_header(ws, RECETAS_COLS)
    for rec in session.scalars(select(Recipe).order_by(Recipe.id)).all():
        ws.append([rec.id, rec.name, rec.yield_qty, rec.yield_unit, rec.notes])
    _autosize(ws)

    # --- Lineas (with denormalized names for human readability + import) ---
    ws = wb.create_sheet("Lineas")
    _write_header(ws, LINEAS_COLS)
    ingredients_by_id = {ing.id: ing for ing in session.scalars(select(Ingredient)).all()}
    recipes_by_id = {rec.id: rec for rec in session.scalars(select(Recipe)).all()}
    for recipe in recipes_by_id.values():
        for line in recipe.lines:
            if line.line_kind == "ingredient":
                target_name = ingredients_by_id.get(line.line_ref_id)
                target_name_str = target_name.name if target_name else None
            elif line.line_kind == "sub_recipe":
                target_name = recipes_by_id.get(line.line_ref_id)
                target_name_str = target_name.name if target_name else None
            else:
                target_name_str = None
            ws.append(
                [
                    line.id,
                    line.recipe_id,
                    recipe.name,
                    line.line_kind,
                    line.line_ref_id,
                    target_name_str,
                    line.qty,
                    line.notes,
                ]
            )
    _autosize(ws)

    # --- Productos ---
    ws = wb.create_sheet("Productos")
    _write_header(ws, PRODUCTOS_COLS)
    for prod in session.scalars(select(Product).order_by(Product.id)).all():
        recipe_name = recipes_by_id.get(prod.recipe_id)
        ws.append(
            [
                prod.id,
                prod.name,
                prod.portion_label,
                _money_cell(prod.sale_price_gs),
                prod.recipe_id,
                recipe_name.name if recipe_name else None,
                prod.notes,
            ]
        )
    _autosize(ws)

    # --- Ventas ---
    ws = wb.create_sheet("Ventas")
    _write_header(ws, VENTAS_COLS)
    products_by_id = {prod.id: prod for prod in session.scalars(select(Product)).all()}
    for sale in session.scalars(select(Sale).order_by(Sale.id)).all():
        product = products_by_id.get(sale.product_id)
        ws.append(
            [
                sale.id,
                sale.sold_at,
                sale.product_id,
                product.name if product else None,
                sale.qty,
                _money_cell(sale.unit_price_gs),
                sale.notes,
                sale.voided_at,
            ]
        )
    _autosize(ws)

    # --- StockMoves (derived from sales; informational only) ---
    ws = wb.create_sheet("StockMoves")
    _write_header(ws, STOCKMOVES_COLS)
    for sale in session.scalars(select(Sale).order_by(Sale.id)).all():
        for move in sale.stock_moves:
            ws.append(
                [
                    move.id,
                    move.sale_id,
                    move.affected_recipe_id,
                    move.ingredient_id,
                    move.qty_delta,
                ]
            )
    _autosize(ws)

    wb.save(str(path))
    return path.resolve()


def to_bytes(session: Session) -> bytes:
    """Same as to_file but returns bytes (for streaming downloads).

    Uses openpyxl's BytesIO-friendly save.
    """
    from io import BytesIO

    path = BytesIO()
    # We can save to BytesIO directly with openpyxl
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    ws = wb.create_sheet("Ingredientes")
    _write_header(ws, INGREDIENTES_COLS)
    for ing in session.scalars(select(Ingredient).order_by(Ingredient.id)).all():
        ws.append(
            [
                ing.id,
                ing.name,
                ing.unit,
                ing.stock_qty,
                _money_cell(ing.purchase_price_gs),
                ing.min_stock_qty,
                ing.notes,
            ]
        )

    ws = wb.create_sheet("Recetas")
    _write_header(ws, RECETAS_COLS)
    for rec in session.scalars(select(Recipe).order_by(Recipe.id)).all():
        ws.append([rec.id, rec.name, rec.yield_qty, rec.yield_unit, rec.notes])

    ws = wb.create_sheet("Lineas")
    _write_header(ws, LINEAS_COLS)
    ingredients_by_id = {ing.id: ing for ing in session.scalars(select(Ingredient)).all()}
    recipes_by_id = {rec.id: rec for rec in session.scalars(select(Recipe)).all()}
    for recipe in recipes_by_id.values():
        for line in recipe.lines:
            if line.line_kind == "ingredient":
                target_name = ingredients_by_id.get(line.line_ref_id)
                target_name_str = target_name.name if target_name else None
            elif line.line_kind == "sub_recipe":
                target_name = recipes_by_id.get(line.line_ref_id)
                target_name_str = target_name.name if target_name else None
            else:
                target_name_str = None
            ws.append(
                [
                    line.id,
                    line.recipe_id,
                    recipe.name,
                    line.line_kind,
                    line.line_ref_id,
                    target_name_str,
                    line.qty,
                    line.notes,
                ]
            )

    ws = wb.create_sheet("Productos")
    _write_header(ws, PRODUCTOS_COLS)
    for prod in session.scalars(select(Product).order_by(Product.id)).all():
        recipe_name = recipes_by_id.get(prod.recipe_id)
        ws.append(
            [
                prod.id,
                prod.name,
                prod.portion_label,
                _money_cell(prod.sale_price_gs),
                prod.recipe_id,
                recipe_name.name if recipe_name else None,
                prod.notes,
            ]
        )

    ws = wb.create_sheet("Ventas")
    _write_header(ws, VENTAS_COLS)
    products_by_id = {prod.id: prod for prod in session.scalars(select(Product)).all()}
    for sale in session.scalars(select(Sale).order_by(Sale.id)).all():
        product = products_by_id.get(sale.product_id)
        ws.append(
            [
                sale.id,
                sale.sold_at,
                sale.product_id,
                product.name if product else None,
                sale.qty,
                _money_cell(sale.unit_price_gs),
                sale.notes,
                sale.voided_at,
            ]
        )

    ws = wb.create_sheet("StockMoves")
    _write_header(ws, STOCKMOVES_COLS)
    for sale in session.scalars(select(Sale).order_by(Sale.id)).all():
        for move in sale.stock_moves:
            ws.append(
                [
                    move.id,
                    move.sale_id,
                    move.affected_recipe_id,
                    move.ingredient_id,
                    move.qty_delta,
                ]
            )

    wb.save(path)
    return path.getvalue()


__all__ = ["to_file", "to_bytes"]

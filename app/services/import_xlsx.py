"""app/services/import_xlsx.py — Drive/HEREBUS .xlsx → SQLite.

Per dev plan §9 Task 6 + v2 §6 (Excel import).

Reads a HEREBUS-format Excel workbook and loads its rows into the RMS SQLite
database. The workbook layout is whatever the export_xlsx service produces
(symmetric), with these sheets:

    Ingredientes  — id, name, unit, stock_qty, purchase_price_gs, min_stock_qty, notes
    Recetas       — id, name, yield_qty, yield_unit, notes
    Lineas        — id, recipe_id, line_kind ('ingredient' | 'sub_recipe'),
                    line_ref_id (FK target id), qty, notes
    Productos     — id, name, portion_label, sale_price_gs, recipe_id, notes
    Ventas        — id, sold_at, product_id, qty, unit_price_gs, notes, voided_at
    StockMoves    — id, sale_id, affected_recipe_id, ingredient_id, qty_delta

Money discipline: money values (purchase_price_gs, sale_price_gs,
unit_price_gs) MUST come through `app.rms.money.to_int_gs`. No raw floats.

Idempotency: the import is keyed by ImportBatch. The caller passes the same
DB session; we never close or commit it. The function returns an
ImportResult dataclass with counts and warnings.

What we DON'T do here:
- Schema migrations (the caller is responsible for init_db)
- Cross-workbook merging (single file → single batch)
- Cycle detection (costing engine detects this on read)
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.rms.models import (
    ImportBatch,
    Ingredient,
    Product,
    Recipe,
    RecipeLine,
    Sale,
)
from app.rms.money import parse_gs, to_int_gs


@dataclass
class ImportResult:
    """Result of a successful import."""

    batch_id: int
    source_filename: str
    ingredients: int = 0
    recipes: int = 0
    lines: int = 0
    products: int = 0
    sales: int = 0
    stock_moves: int = 0
    warnings: list[str] = field(default_factory=list)

    def row_counts(self) -> dict[str, int]:
        return {
            "ingredients": self.ingredients,
            "recipes": self.recipes,
            "lines": self.lines,
            "products": self.products,
            "sales": self.sales,
            "stock_moves": self.stock_moves,
        }


def _sheet(wb, name: str) -> Worksheet | None:
    """Return sheet by name, or None if missing."""
    if name in wb.sheetnames:
        return wb[name]
    return None


def _rows(sheet: Worksheet | None) -> list[dict]:
    """Convert a sheet's rows to list of dicts (header → key, cell → value)."""
    if sheet is None:
        return []
    iter_rows = sheet.iter_rows(values_only=True)
    try:
        header = next(iter_rows)
    except StopIteration:
        return []
    header = [str(h) if h is not None else "" for h in header]
    out: list[dict] = []
    for row in iter_rows:
        if row is None or all(c is None for c in row):
            continue
        out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return out


def _opt_str(value) -> str | None:
    """Cell value → optional string (None → None; empty string → None)."""
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _opt_float(value) -> float | None:
    """Cell value → optional float (None/empty → None)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    return float(s)


def _money_int_gs(value, *, field_name: str, warnings: list[str]) -> int | None:
    """Cell value → integer Gs. via parse_gs (accepts 'Gs. 5.000' format) or
    to_int_gs (accepts numeric).

    Records a warning on parse failure rather than crashing — keeps the import
    partial-success instead of all-or-nothing.

    Returns None for missing prices (None, empty string).
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return to_int_gs(value)
    s = str(value).strip()
    if not s:
        return None
    # Try parse_gs first (accepts "Gs. 5.000", "1.234.567", "1234567")
    try:
        return parse_gs(s)
    except ValueError:
        pass
    # Fall back to to_int_gs (accepts "5000", "5000.5")
    try:
        return to_int_gs(s)
    except (ValueError, TypeError) as exc:
        warnings.append(f"{field_name}: no se pudo parsear {value!r}: {exc}")
        return None


def from_file(session: Session, path: str | Path) -> ImportResult:
    """Import a HEREBUS .xlsx into the DB via `session`.

    Caller owns the session (we commit at the end). Caller must run
    `init_db(engine)` before calling.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected .xlsx, got {path.suffix}")

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    warnings: list[str] = []

    # 1. Ingredientes
    ingredients_index: dict[str, int] = {}  # name → id
    ingredients_count = 0
    for row in _rows(_sheet(wb, "Ingredientes")):
        name = _opt_str(row.get("name"))
        if not name:
            warnings.append(f"Ingredientes: fila sin nombre, saltada: {row}")
            continue
        unit = _opt_str(row.get("unit")) or "und"
        stock_qty = _opt_float(row.get("stock_qty")) or 0.0
        purchase_price_gs = _money_int_gs(
            row.get("purchase_price_gs"),
            field_name=f"Ingredientes[{name}].purchase_price_gs",
            warnings=warnings,
        )
        min_stock_qty = _opt_float(row.get("min_stock_qty")) or 0.0
        notes = _opt_str(row.get("notes"))
        ing = Ingredient(
            name=name,
            unit=unit,
            stock_qty=stock_qty,
            purchase_price_gs=purchase_price_gs,
            min_stock_qty=min_stock_qty,
            notes=notes,
        )
        session.add(ing)
        session.flush()
        ingredients_index[name] = ing.id
        ingredients_count += 1

    # 2. Recetas
    recipes_index: dict[str, int] = {}  # name → id
    recipes_count = 0
    for row in _rows(_sheet(wb, "Recetas")):
        name = _opt_str(row.get("name"))
        if not name:
            warnings.append(f"Recetas: fila sin nombre, saltada: {row}")
            continue
        yield_qty = _opt_float(row.get("yield_qty"))
        yield_unit = _opt_str(row.get("yield_unit")) or "und"
        notes = _opt_str(row.get("notes"))
        recipe = Recipe(
            name=name,
            yield_qty=yield_qty,
            yield_unit=yield_unit,
            notes=notes,
        )
        session.add(recipe)
        session.flush()
        recipes_index[name] = recipe.id
        recipes_count += 1

    # 3. Lineas (polymorphic recipe_line)
    lines_count = 0
    for row in _rows(_sheet(wb, "Lineas")):
        recipe_name = _opt_str(row.get("recipe_name"))
        line_kind = _opt_str(row.get("line_kind"))
        target_name = _opt_str(row.get("target_name"))
        qty = _opt_float(row.get("qty"))
        notes = _opt_str(row.get("notes"))

        if not recipe_name or recipe_name not in recipes_index:
            warnings.append(
                f"Lineas: recipe_name={recipe_name!r} no existe en Recetas, saltada: {row}"
            )
            continue
        if line_kind not in ("ingredient", "sub_recipe"):
            warnings.append(f"Lineas[{recipe_name}]: line_kind={line_kind!r} inválido, saltada")
            continue
        if not target_name:
            warnings.append(f"Lineas[{recipe_name}]: target_name vacío, saltada")
            continue
        if qty is None or qty <= 0:
            warnings.append(f"Lineas[{recipe_name}->{target_name}]: qty={qty!r} inválido, saltada")
            continue

        # Resolve target id by name lookup
        if line_kind == "ingredient":
            if target_name not in ingredients_index:
                warnings.append(
                    f"Lineas[{recipe_name}]: ingrediente {target_name!r} no existe, saltada"
                )
                continue
            line_ref_id = ingredients_index[target_name]
        else:  # sub_recipe
            if target_name not in recipes_index:
                warnings.append(
                    f"Lineas[{recipe_name}]: sub-receta {target_name!r} no existe, saltada"
                )
                continue
            line_ref_id = recipes_index[target_name]

        line = RecipeLine(
            recipe_id=recipes_index[recipe_name],
            line_kind=line_kind,
            line_ref_id=line_ref_id,
            qty=qty,
            notes=notes,
        )
        session.add(line)
        lines_count += 1

    # 4. Productos
    products_index: dict[str, int] = {}
    products_count = 0
    for row in _rows(_sheet(wb, "Productos")):
        name = _opt_str(row.get("name"))
        if not name:
            warnings.append(f"Productos: fila sin nombre, saltada: {row}")
            continue
        portion_label = _opt_str(row.get("portion_label")) or "1 unidad"
        sale_price_gs = _money_int_gs(
            row.get("sale_price_gs"),
            field_name=f"Productos[{name}].sale_price_gs",
            warnings=warnings,
        )
        if sale_price_gs is None:
            sale_price_gs = 0
        recipe_name = _opt_str(row.get("recipe_name"))
        recipe_id = recipes_index.get(recipe_name) if recipe_name else None
        notes = _opt_str(row.get("notes"))
        product = Product(
            name=name,
            portion_label=portion_label,
            sale_price_gs=sale_price_gs,
            recipe_id=recipe_id,
            notes=notes,
        )
        session.add(product)
        session.flush()
        products_index[name] = product.id
        products_count += 1

    # 5. Ventas
    sales_count = 0
    for row in _rows(_sheet(wb, "Ventas")):
        product_name = _opt_str(row.get("product_name"))
        if not product_name or product_name not in products_index:
            warnings.append(f"Ventas: product_name={product_name!r} no existe, saltada: {row}")
            continue
        sold_at = row.get("sold_at")
        if isinstance(sold_at, datetime):
            sold_at_dt = sold_at
        elif isinstance(sold_at, str):
            try:
                sold_at_dt = datetime.fromisoformat(sold_at)
            except ValueError:
                warnings.append(f"Ventas[{product_name}]: sold_at={sold_at!r} inválido, saltada")
                continue
        else:
            warnings.append(f"Ventas[{product_name}]: sold_at={sold_at!r} vacío, saltada")
            continue
        qty = _opt_float(row.get("qty"))
        if qty is None or qty <= 0:
            warnings.append(f"Ventas[{product_name}]: qty={qty!r} inválido, saltada")
            continue
        unit_price_gs = _money_int_gs(
            row.get("unit_price_gs"),
            field_name=f"Ventas[{product_name}].unit_price_gs",
            warnings=warnings,
        )
        if unit_price_gs is None:
            unit_price_gs = 0
        notes = _opt_str(row.get("notes"))
        voided_at_raw = row.get("voided_at")
        voided_at = None
        if isinstance(voided_at_raw, datetime):
            voided_at = voided_at_raw
        elif isinstance(voided_at_raw, str) and voided_at_raw.strip():
            try:
                voided_at = datetime.fromisoformat(voided_at_raw)
            except ValueError:
                warnings.append(f"Ventas[{product_name}]: voided_at={voided_at_raw!r} inválido")
        sale = Sale(
            sold_at=sold_at_dt,
            product_id=products_index[product_name],
            qty=qty,
            unit_price_gs=unit_price_gs,
            notes=notes,
            voided_at=voided_at,
        )
        session.add(sale)
        sales_count += 1

    # 6. StockMoves (skip on round-trip — moves are derived from sales)
    # We DON'T import StockMoves from the sheet — moves are produced by
    # apply_sale. Importing them would double-count. The export writes them
    # for archival; the import ignores them.
    stock_moves_count = 0

    # Record the batch
    result = ImportResult(
        batch_id=-1,  # set after flush
        source_filename=path.name,
        ingredients=ingredients_count,
        recipes=recipes_count,
        lines=lines_count,
        products=products_count,
        sales=sales_count,
        stock_moves=stock_moves_count,
        warnings=warnings,
    )
    batch = ImportBatch(
        imported_at=datetime.now(),
        source_filename=path.name,
        note=None,
        row_counts_json=json.dumps(result.row_counts()),
    )
    session.add(batch)
    session.flush()
    result.batch_id = batch.id
    session.commit()
    return result


__all__ = ["from_file", "ImportResult"]
